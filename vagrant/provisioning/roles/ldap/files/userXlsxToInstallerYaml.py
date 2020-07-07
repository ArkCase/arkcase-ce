#!/usr/bin/env python
import sys
import openpyxl

if len(sys.argv) < 3:
    print("Must provide Excel input file name and starting row")
    sys.exit(1)

excel_file = sys.argv[1]
start_row = int(sys.argv[2])

wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

def fnFirstInitLnFirstX(firstName, lastName, numLnChars):
    """Construct a samAccountName by first char of first name, up to seven chars 
    of last name"""
    return (firstName[0:1] + lastName[0:numLnChars]).lower()

def getCellValue(sheet, row, column):
    """Get value of a cell at the given index"""
    return sheet[openpyxl.utils.cell.get_column_letter(column) + str(row)].value

def formatGroupProperty(group):
    """Format a group name with the proper ldap_prefix and ldap_group_base placeholders"""
    return '      - "CN={{ ldap_prefix }}' + group[0:group.find('@')] + ',{{ldap_group_base}}"'

for r in range(start_row, sheet.max_row + 1):
    firstName = getCellValue(sheet, r, 1)
    lastName = getCellValue(sheet, r, 2)
    email = getCellValue(sheet, r, 3)
    groups = getCellValue(sheet, r, 6)
    office = getCellValue(sheet, r, 8)

    # see if there is really a value here
    if ( hasattr(firstName, 'strip')):
        print('  - user_id: "{{ ldap_prefix }}' + fnFirstInitLnFirstX(firstName, lastName, 6) + '"')
        print("    description: " + firstName + " " + lastName + " - " + office)
        print("    name: " + firstName + " " + lastName)
        print("    mail: " + email)
        print("    userPrincipalName: " + email.lower())
        print("    firstName: " + firstName)
        print("    lastName: " + lastName)
        print('    password: "{{ default_user_password }}"')
        print("    groups:")

        for group in groups.split():
            if group.find("@") > 0:
                print(formatGroupProperty(group))
        
        print("    nonexpiring_password: yes")
   
wb.close()
