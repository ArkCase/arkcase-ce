#!/usr/bin/env python
import sys
import openpyxl

if len(sys.argv) < 4:
    print "Must provide Excel file name, and at least one old_string,new_string pair"
    sys.exit(1)

excel_file = sys.argv[1] + ".xlsx"
new_excel_file = sys.argv[1] + "-updated.xlsx"

wb = openpyxl.load_workbook(excel_file)
sheet = wb.get_sheet_by_name('Sheet1')

updated = False

for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        rowcol_index = openpyxl.utils.cell.get_column_letter(c) + str(r)
        v = sheet[rowcol_index].value
        for p in range(2, len(sys.argv), 2):
            old_string = sys.argv[p]
            new_string = sys.argv[p + 1]
            if old_string <> new_string and hasattr(v, 'strip') and v.strip() == old_string:
                print "replacing " + v
                sheet[rowcol_index] = new_string
                updated = True

if updated:
    wb.save(filename = new_excel_file)
    
wb.close()
